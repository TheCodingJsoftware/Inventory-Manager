import contextlib
from datetime import date
from itertools import zip_longest

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rich import print

from utils.workspace.item import Item


class MondayExcelFile:
    """Getting data from an excel file generated by monday.com."""

    def __init__(self, path: str) -> None:
        """
        The function initializes an object with a given path and sets up attributes for columns to find
        and data storage.

        Args:
          path (str): The `path` parameter is a string that represents the file path where the data is
        stored.
        """
        self.path: str = path
        self.columns_to_find: list[str] = ["name", "paint color", "thickness", "material type", "parts per"]  # in lower
        self.data: dict[str, dict[dict[str, int], dict[str, any]]] = {}

    def find_jobs(self, sheet: Worksheet) -> None:
        """
        The function iterates through rows in a worksheet, checks if a cell value matches certain
        conditions, and adds data to a dictionary.

        Args:
          sheet (Worksheet): The parameter "sheet" is of type Worksheet. It represents the Excel
        worksheet that contains the data.
        """
        for row_index in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=1)
            cell_value = cell.value
            if not cell_value or not isinstance(cell_value, str):
                continue
            if cell_value.lower() in self.columns_to_find and cell_value.lower() == "name":
                cell_above = sheet.cell(row=cell.row - 1, column=cell.column)  # Job name
                self.data[cell_above.value] = {"job_data": {"start_row": cell.row}, "items": {}}

    def find_table_headers(self, sheet: Worksheet) -> None:
        """
        The function `find_table_headers` iterates through a given worksheet and identifies table
        headers based on specified criteria, storing the corresponding data in a dictionary.

        Args:
          sheet (Worksheet): The parameter `sheet` is of type `Worksheet`.
        """
        for job_index, (job_name, job_data) in enumerate(self.data.items()):
            job_row: int = job_data["job_data"]["start_row"]
            try:
                end_row = self.data[list(self.data.keys())[job_index + 1]]["job_data"]["start_row"] - 3  # needs to offset to the next job name
            except IndexError:
                end_row = sheet.max_row
            for col_index in range(1, sheet.max_column):
                header_cell = sheet.cell(row=job_row, column=col_index)
                if header_cell.value.lower() in self.columns_to_find:
                    job_data["items"][header_cell.value.lower()] = []
                    for row in range(job_row + 1, end_row):
                        cell_item = sheet.cell(row=row, column=col_index)
                        job_data["items"][header_cell.value.lower()].append(cell_item.value)

    def clean_up_data(self) -> dict[str, dict[str, any]]:
        """
        The `clean_up_data` function takes in a dictionary of job data, iterates through the items in
        each job, and creates a new dictionary with cleaned up data by removing any items that don't
        meet certain conditions.

        Returns:
          the result of calling the `remove_empty_jobs` method on the `self` object, passing in the
        `data` dictionary as an argument.
        """
        data: dict[str, dict[str, any]] = {}
        for job_name, job_data in self.data.items():
            data[job_name] = {}
            # for headers, item_data in job_data["items"].items():
            for item_name, paint_color, thickness, material_type, parts_per in zip_longest(
                job_data["items"]["name"],
                job_data["items"]["paint color"],
                job_data["items"]["thickness"],
                job_data["items"]["material type"],
                job_data["items"]["parts per"],
                fillvalue=None,
            ):
                if item_name is None:
                    continue
                # Its a formula, it should be float or int
                if isinstance(parts_per, str) or item_name == "Subitems" or (thickness == material_type == None):
                    continue
                data[job_name][item_name] = {
                    "paint_color": paint_color,
                    "thickness": thickness,
                    "material_type": material_type,
                    "parts_per": parts_per,
                }
        return self.remove_empty_jobs(data)

    def remove_empty_jobs(self, data: dict[str, dict[str, any]]) -> dict[str, dict[str, any]]:
        """
        The function removes empty jobs from a dictionary of job data.

        Args:
          data (dict[str, dict[str, any]]): A dictionary where the keys are job names (strings) and the
        values are dictionaries containing job data (also dictionaries).

        Returns:
          a new dictionary `new_data` that contains only the non-empty job data from the input
        dictionary `data`.
        """
        new_data: dict[str, dict[str, any]] = {job_name: job_data for job_name, job_data in data.items() if job_data}
        return new_data

    def combine_data(self, data: dict[str, dict[str, any]]) -> dict[str, list[Item]]:
        """
        The function `combine_data` takes a dictionary of dictionaries as input, creates a new
        dictionary with the same keys but with values as lists of `Item` objects, and returns the new
        dictionary.

        Args:
          data (dict[str, dict[str, any]]): The `data` parameter is a dictionary where the keys are job
        names and the values are dictionaries containing item data. Each item data dictionary contains
        information about an item, such as its thickness, material type, paint color, parts per, etc.

        Returns:
          The function `combine_data` returns a dictionary where the keys are strings and the values are
        lists of `Item` objects.
        """
        new_data: dict[str, list[Item]] = {}
        for job_name, job_data in data.items():
            new_data[job_name] = []
            for item_name, item_data in job_data.items():
                item: Item = Item(
                    name=item_name,
                    data={
                        "Bending Files": [],
                        "Welding Files": [],
                        "CNC/Milling Files": [],
                        "thickness": item_data["thickness"],
                        "material": item_data["material_type"],
                        "paint_type": None,
                        "paint_color": item_data["paint_color"],
                        "parts_per": item_data["parts_per"],
                        "flow_tag": [],
                        "timers": {},
                        "customer": "",
                        "ship_to": "",
                        "show": True,
                    },
                )
                new_data[job_name].append(item)
        return new_data

    def get_data(self) -> dict[str, list[Item]]:
        """
        The function `get_data` loads data from an Excel workbook, finds jobs and table headers in the
        sheet, and combines and cleans up the data before returning it.

        Returns:
          a dictionary with string keys and list values.
        """
        workbook = openpyxl.load_workbook(filename=self.path)
        sheet = workbook.active
        self.find_jobs(sheet=sheet)
        self.find_table_headers(sheet=sheet)
        return self.combine_data(self.clean_up_data())
